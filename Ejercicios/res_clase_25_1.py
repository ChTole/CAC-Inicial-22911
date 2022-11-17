"""
Dificultad adicional 1: generar un nuevo script de Python que realice una búsqueda dentro del libro creado permitiendo seleccionar la columna a filtrar.
"""

from openpyxl import load_workbook

try:
    ruta = r'D:\CodoACodo\Comision 22911\Clase32\datosJson.xlsx'
    libro = load_workbook(filename= ruta)
except FileNotFoundError:
    print('No encontré el archivo!')
    ruta = input('Ingrese ubicación correcta del archivo')
    libro = load_workbook(filename= ruta)
except:
    print('Ocurrió un error inesperado!')
    
# columna = libro['DatosJSON'][1][0] # accedo a la primera celda de la primera fila  "A1"
# print(columna.value) # accedo al valor que contiene la celda "A1".value

def buscar_en_libro(opcion):
    resultado = False
    opcion = int(opcion)
    columna = libro['DatosJSON'][1][opcion]
    busqueda = input(f'Ingrese {columna.value} a buscar: ')
    # Tarea: evaluar la variable busqueda para que no se interrumpa la ejecución.
    for fila in libro['DatosJSON'].iter_rows(min_row=2, max_row=libro['DatosJSON'].max_row):
        # if busqueda in fila[opcion].value:
        if busqueda.lower() in str(fila[opcion].value).lower():
        # permitir evaluar mayúsculas y minúsculas
            print('Se encontró:')
            print(f'Se encontró {fila[0].value} {fila[1].value} {fila[2].value} {fila[3].value} {fila[4].value}')
            resultado = True
    if not resultado:
        print('No encontré lo que buscabas!')

opcion = ''
while opcion != '5':
    print("""
    Opciones de búsqueda:
    0 - Por Id
    1 - Por Nombre
    2 - Por Correo
    3 - Por Teléfono
    4 - Por Sitio web
    5 - Salir del menú          
          """)
    opcion = input('Ingrese su opción: ')
    # Tarea: evaluar la variable opcion para que no se interrumpa la ejecución.
    try:
        if int(opcion) in range(5):      
            buscar_en_libro(opcion)
        elif opcion == '5':
            print('Gracias por usar la app! Hasta pronto!')
            continue
        else:
            print('Opción inválida!!!')
            opcion = input('Ingrese nuevamente su opción: ')
    except ValueError:
        print('La opción ingresada es incorrecta!')
        continue
libro.close()

    
    