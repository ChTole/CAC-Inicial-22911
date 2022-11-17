"""
Dificultad adicional 2: generar un nuevo script de Python que permita agregar una columna llamada "fecha de nacimiento" y solicite por consola el ingreso de los datos para cada registro. Utilizar el módulo datetime y corroborar que cada dato ingresado para construir la fecha sea de acuerdo al formato dd/mm/aaaa. Por ejemplo:
- si ingreso 23/04/1998, la fecha se crea
- si ingreso 43/15/2005, la fecha no se crea.
"""

from datetime import date
from openpyxl import load_workbook

ruta = r'D:\CodoACodo\Comision 22911\Clase32\datosJson.xlsx'
libro = load_workbook(filename=ruta)
hoja = libro["DatosJSON"]
hoja["f1"] = "Fecha de nacimiento"

for fila in hoja.iter_rows(min_row=2, max_row=hoja.max_row):
    fecha_nac = input(f"Ingrese fecha de nacimiento para {fila[1].value} (dd/mm/aaaa): ")
    while True:
        try:
            fila[5].value = date(
            int(fecha_nac.split('/')[2]), 
            int(fecha_nac.split('/')[1]), 
            int(fecha_nac.split('/')[0]), 
            )
            break
        except Exception as e:
            print(type(e).__name__)
            print('Fecha inválida!')
            fecha_nac = input(f"Ingrese fecha de nacimiento para {fila[1].value} válida (dd/mm/aaaa): ")

try:
    libro.save(filename=ruta)
except PermissionError:
    print("""
    Por favor cerrá el libro para guardarlo,
    cuando lo hagas, pulsa ENTER.
    """)
    input(">>>")
    libro.save(filename=ruta)
        
libro.close()